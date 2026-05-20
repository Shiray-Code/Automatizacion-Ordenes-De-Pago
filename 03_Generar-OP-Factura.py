import os
import json
import shutil

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# =====================================================
# RUTAS
# =====================================================

JSON_FACTURAS = "facturas.json"

JSON_ODOO = "odoo_data.json"

PASTA_TEMP = r"(Datos Personales)"

PASTA_OP = r"(Datos Personales)"

PLANTILLA = r"(Datos Personales)"

CORRELATIVO_FILE = r"(Datos Personales)"

# =====================================================
# LEER JSON
# =====================================================

with open(
    JSON_FACTURAS,
    "r",
    encoding="utf-8"
) as f:

    facturas_por_proveedor = json.load(f)

with open(
    JSON_ODOO,
    "r",
    encoding="utf-8"
) as f:

    odoo_data = json.load(f)

# =====================================================
# CORRELATIVO
# =====================================================

with open(
    CORRELATIVO_FILE,
    "r"
) as f:

    correlativo_actual = int(
        f.read().strip()
    )

# =====================================================
# PROVEEDORES
# =====================================================

ops_generadas = []

OPS_GENERADAS_JSON = r"(Datos Personales)"

for proveedor, facturas in facturas_por_proveedor.items():

    correlativo_actual += 1

    with open(
        CORRELATIVO_FILE,
        "w"
    ) as f:

        f.write(
            str(correlativo_actual)
        )

    nombre_op = (
        f"OP-KS-00{correlativo_actual}"
    )

    carpeta_proveedor = os.path.join(
        PASTA_OP,
        proveedor
    )

    os.makedirs(
        carpeta_proveedor,
        exist_ok=True
    )

    carpeta_op = os.path.join(
        carpeta_proveedor,
        nombre_op
    )

    os.makedirs(
        carpeta_op,
        exist_ok=True
    )

    # =====================================================
    # MOVER TEMP
    # =====================================================

    for factura in facturas:

        oc = factura["oc"]

        carpeta_temp_oc = os.path.join(
            PASTA_TEMP,
            oc
        )

        if os.path.exists(
            carpeta_temp_oc
        ):

            for archivo in os.listdir(
                carpeta_temp_oc
            ):

                ruta_origen = os.path.join(
                    carpeta_temp_oc,
                    archivo
                )

                ruta_destino = os.path.join(
                    carpeta_op,
                    archivo
                )

                try:

                    shutil.move(
                        ruta_origen,
                        ruta_destino
                    )

                except:

                    pass

    # =====================================================
    # EXCEL
    # =====================================================

    wb = load_workbook(
        PLANTILLA
    )

    ws = wb[
        "Requerimiento (2)"
    ]

    for fila in ws["B16:P31"]:

        for celda in fila:

            if not isinstance(
                celda,
                MergedCell
            ):

                celda.value = None

    fila_actual = 17

    proyectos = set()

    centros = set()

    descripciones = []

    # =====================================================
    # FACTURAS
    # =====================================================

    for factura in facturas:

        oc = factura["oc"]

        if oc in odoo_data:

            proyecto = odoo_data[
                oc
            ]["proyecto"]

            centro = odoo_data[
                oc
            ]["centro_costo"]

            if proyecto:

                proyectos.add(
                    proyecto
                )

            if centro:

                centros.add(
                    centro
                )

        ws[
            f"B{fila_actual}"
        ] = f"OC {factura['oc']}"

        ws[
            f"B{fila_actual+1}"
        ] = f"HES {factura['hes']}"

        ws[
            f"B{fila_actual+2}"
        ] = f"FACT. {factura['factura']}"

        ws[
            f"L{fila_actual+2}"
        ] = factura["monto"]

        descripciones.append(
            factura["descripcion"]
        )

        try:

            shutil.copy2(

                factura["ruta_pdf"],

                os.path.join(
                    carpeta_op,
                    factura["archivo"]
                )
            )

        except:

            pass

        fila_actual += 4

    ws["D4"] = proveedor

    ws["D1"] = "\n".join(
        sorted(proyectos)
    )

    ws["B33"] = "\n".join(
        sorted(centros)
    )

    ws["B15"] = "\n".join(
        list(set(descripciones))
    )

    ws["S1"] = correlativo_actual

    ruta_excel = os.path.join(

        carpeta_op,
        f"{nombre_op}.xlsx"
    )

    wb.save(
        ruta_excel
    )

    print(f"OP creada: {nombre_op}")
    
    ops_generadas.append({

    "nombre_op": nombre_op,
    "ruta": carpeta_op

})

with open(
    OPS_GENERADAS_JSON,
    "w",
    encoding="utf-8"
) as f:

    json.dump(
        ops_generadas,
        f,
        indent=4,
        ensure_ascii=False
    )

print("\nPROCESO TERMINADO")
